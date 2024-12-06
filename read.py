from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import StringFilter
import pandas as pd
import re
"""
This contains two important functions
search_single_value and search_in_workbook
search_single_value will search for a single value in a column and return the cell location
search_in_workbook will search for multiple values in a column and return the cell location, including conditions as 
well as printing their associated values
"""
# Load the Excel file
book = load_workbook("FoundHouse.xlsx")  # Load the excel file
def search_single_value(sheet_name, column_header, target):
    """
    Search for a single value in a column and return the cell location.
    sheet_name : The name of the sheet to search in the excel file
    column_header : The column header to search in the sheet
    target : The value to search for in the column
    """
    result = ""
    if sheet_name and column_header and target:  # Check that all the inputs are not empty
        sheet = book[sheet_name]  # Get the sheet name from the excel workbook
        headers = []  # This will serve to store the column headers
        for cell in sheet[2] or sheet[0]:  # Read the first row to get column headers
            headers.append(cell.value)  # Add the column header to the list
        if column_header.isalpha() and len(column_header) == 1:  # Check if the column header is a single letter
            column_letter = column_header.upper()  # Match the column letter to the column header
        else:
            column_letter = chr(ord('A') + headers.index(column_header))
        
        column = sheet[column_letter]  # Get the column from the sheet
        if target.isdigit() and int(target) > 0:  # Check if the target is an integer
            target = int(target)
        for cell in column:  # Loop through the column
            if str(cell.value).casefold().strip() == str(target).casefold().strip() or cell.value == target:
                # Return the cell location if target is found
                return f'Found {target} in cell {column_letter}{cell.row}'
            row = sheet[cell.row]  # Get the row
            for cell in row:  # Loop through the row
                result += str(cell.value) + "\t"  # Print the value of the cell
            result += "\n"
    return result
def check_conditions(cell_value, operator, value):
    """
    this exist to check the conditions
    example: looking for everything that has an age of 5 or more
    """
    if operator == '=':
        return str(cell_value).casefold().strip() == str(value).casefold().strip()
    elif operator == '<=':
        return cell_value <= int(value)
    elif operator == '>=':
        return cell_value >= int(value)
    elif operator == '<':
        return cell_value < int(value)
    elif operator == '>':
        return cell_value > int(value)
    else:
        return False

def search_in_workbook(sheet_name, targets):
    """
    Allows for filtering by multiple conditions
    Able to search multiple things at once such as Dog, Cat, and Bird. It will return the cell location
    as well as print their associated values
    """
    conditions = []  # To store the conditions
    for target in targets:  # Loop through the targets
        # Check if the target is a condition
        match = re.match(r'(\w+)\s*([<>=]+)\s*(\w+)', target)
        if match is not None:  # If it is a condition, parse it
            column, operator, value = match.groups()
            conditions.append((column, operator, value))
        else:
            # If the target is not a condition, treat it as a single value
            conditions.append((None, None, target))

    sheet = book[sheet_name]  # Get the sheet name
    headers = [cell.value for cell in sheet[2]]  # Read the first row to get column headers
    result = ""  # To store the results

    for column_index, column in enumerate(sheet.columns):  # Loop through the columns
        column_header = headers[column_index]  # Grabbing the column header
        for cell in column[2:]:
            row_index = cell.row  # Get the row
            row_values = [cell.value for cell in sheet[row_index]]  # Get the values in the row
            match_any_conditions = False
            found_target = None
            for column_name, operator, value in conditions: 
                # Check if the condition is a single value
                if column_name is None:
                    # If the condition is a single value, check if it's present in the row
                    if value in row_values:
                        match_any_conditions = True
                        found_target = value
                        break
                else:
                    # Check if the column header matches the condition
                    if column_name == column_header:
                        cell_value = cell.value  # Get the value of the cell
                        if check_conditions(cell_value, operator, value):
                            match_any_conditions = True
                            found_target = value
                            break

            if match_any_conditions:
                result += f"Found {found_target} in cell {cell.column_letter}{cell.row}\n"
                row = sheet[cell.row]  # Get the row
                for cell in row:
                    result += str(cell.value) + "\t"  # Print the value of the cell
                result += "\n"

    return result
