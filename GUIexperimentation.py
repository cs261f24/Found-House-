import sys
import pandas as pd
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QApplication, QLabel, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, 
    QComboBox, QTableWidget, QTextEdit, QSplitter, QFrame
)
from openpyxl import load_workbook
from read import search_in_workbook, search_single_value


class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()

        # Set the window title and size
        self.setWindowTitle("Found House - For the Pets")
        self.setGeometry(100, 100, 1200, 800)  # Window size and position

        # Create the main vertical layout
        """
        QVBoxLayout is a class in PyQt5 that is used to create a vertical layout. The layout is used to arrange widgets in a vertical manner.
        """
        main_layout = QVBoxLayout(self)

        # Create a horizontal layout for add/remove controls
        """
        QHBoxLayout is a class in PyQt5 that is used to create a horizontal layout. The layout is used to arrange widgets in a horizontal manner.
        """
        add_remove_controls = QHBoxLayout()

        # Input field for entering row values or column name
        """
        QLineEdit is a class in PyQt5 that is used to create a single-line input field. The input field can be used to enter text, numbers, etc.
        """
        self.input_field = QLineEdit() #this will create the input field
        self.input_field.setPlaceholderText("Enter row values (comma-separated) or column name") #this will create the placeholder to serve as a text for the input field
        add_remove_controls.addWidget(self.input_field) #this will add the input field to the add/remove controls layout

        # Button to add rows
        """
        QPushButton is how to create a button in PyQt5
        """
        self.add_rows_button = QPushButton("Add Rows") #create the add rows button
        add_remove_controls.addWidget(self.add_rows_button) 

        # Button to add columns
        self.add_columns_button = QPushButton("Add Columns") 
        add_remove_controls.addWidget(self.add_columns_button)

        # Button to remove rows
        self.remove_rows_button = QPushButton("Remove Rows")
        add_remove_controls.addWidget(self.remove_rows_button)

        # Button to remove columns
        self.remove_columns_button = QPushButton("Remove Columns")
        add_remove_controls.addWidget(self.remove_columns_button)

        # Add the add/remove controls layout to the main layout
        main_layout.addLayout(add_remove_controls)

        # Create a splitter to divide the main layout into left and right sections
        """
        QSplitter is from PyQt5 and it allows us to divide the main layout into left and right sections.
        """
        main_splitter = QSplitter() #this is the horizontal division that will allow the left and right to be side by side
        main_splitter.setOrientation(Qt.Horizontal)  # Horizontal division

        # Left Panel setup

        left_panel = QFrame() #this will create the left panel 
        left_layout = QVBoxLayout() #this will create the layout for the left panel
        left_panel.setLayout(left_layout) #this will set the layout for the left panel

        # Dropdown to select Excel sheets
        """
        QComboBox is from PyQt5 and it helps with creating dropdown menus. The dropdown menu is used to select an item from a list of items.
        """
        self.sheet_input = QComboBox()
        sheets = load_workbook("FoundHouse.xlsx").sheetnames  # Load sheet names from Excel
        self.sheet_input.addItems(sheets)
        left_layout.addWidget(self.sheet_input)

        # Dropdown to select animal species
        self.animal_input_label = QLabel("Select Species:") 
        self.animal_input = QComboBox() #this will create the dropdown menu for the animal species
        self.animal_input.addItems(["Dog", "Cat", "Others"]) # Add options to the dropdown menu
        left_layout.addWidget(self.animal_input_label) #this will add the label to the left layout
        left_layout.addWidget(self.animal_input) #this will add the dropdown menu to the left layout

        # Input field for specifying the column name
        self.column_input = QLineEdit()
        self.column_input.setPlaceholderText("Enter column name")
        left_layout.addWidget(self.column_input)

        # Input field for entering the search target
        self.target_input = QLineEdit()
        self.target_input.setPlaceholderText("Enter what you want to search for(comma-separated for multiple values):")
        left_layout.addWidget(self.target_input)

        # Button to trigger single value search
        self.single_search_button = QPushButton("Search Single Value")
        self.single_search_button.clicked.connect(self.search_single_value_button_clicked)
        left_layout.addWidget(self.single_search_button)

        # Button to trigger search with multiple values
        self.filter_button = QPushButton("Filter by Multiple values")
        self.filter_button.clicked.connect(self.search_in_workbook_button_clicked)
        left_layout.addWidget(self.filter_button)

        # Add the left panel to the splitter
        main_splitter.addWidget(left_panel)

        # Right Panel setup
        right_panel = QFrame()
        right_layout = QVBoxLayout()
        right_panel.setLayout(right_layout)

        # Table to display search results
        self.results_table = QTableWidget()
        right_layout.addWidget(self.results_table)
        
        
        self.input_table = QTableWidget(1, 11)
        self.input_table.setHorizontalHeaderLabels([
            "Name", "Type", "Animal ID", "Pet Name", "Phone",
            "Start Date", "Housed", "Age", "Purpose", "CurrentAge", "Days"
        ]) #this is just setting the column headers(the important thing we want to track)
        self.input_table.horizontalHeader().setStretchLastSection(True) #just allows you to stretch the column headers
        main_layout.addWidget(self.input_table)
        

        # Text edit area to display search results
        self.result_label = QTextEdit()
        right_layout.addWidget(self.result_label)
        
        # Add the right panel to the splitter
        main_splitter.addWidget(right_panel)

        # Add the splitter to the main layout
        main_layout.addWidget(main_splitter)

        # Load the Excel workbook and set the initial sheet
        try:
            self.workbook = load_workbook("FoundHouse.xlsx")
            self.sheet = self.workbook.active  # Set the initial active sheet
        except FileNotFoundError:
            print("Error: Excel file not found. Please ensure the file path is correct.")
            sys.exit()  # Exit the application if the file is not found

        #count animals
        self.count_animals_button = QPushButton("Count Animals")
        self.count_animals_button.clicked.connect(self.count_animals_button_clicked)
        main_layout.addWidget(self.count_animals_button)


    def count_animals_button_clicked(self):
        sheet_name = self.sheet_input.currentText()
        sheetdf = pd.read_excel("FoundHouse.xlsx", sheet_name=sheet_name)
        animal_column = sheetdf.iloc[:, 1]  
        total_animals = len(animal_column) - 1  
        animal_count = animal_column.value_counts() 

        # Display the results 
        result_text = f"The total number of animals in the shelter is: {total_animals}\n\n"
        result_text += "The number of each type of animal is:\n"
        for animal, count in animal_count.items():
            result_text += f"{animal}: {count}\n"
        
        self.result_label.setText(result_text) 

    def search_single_value_button_clicked(self):
        sheet_name = self.sheet_input.currentText() #get the sheet name from the dropdown
        column_letter = self.column_input.text() #get the column name from the input
        target = self.target_input.text() #get the search target from the input
        result = search_single_value(sheet_name, column_letter, target)
        if result:
            self.result_label.setText(result)
        else:
            self.result_label.setText("Target not found")

    def search_in_workbook_button_clicked(self):
        sheet_name = self.sheet_input.currentText()
        targets = self.target_input.text().split(',')
        result = search_in_workbook(sheet_name, targets)
        if result:
            self.result_label.setText(result)
        else:
            self.result_label.setText("Not found")

    def add_rows_button_clicked(self):
        pass

    def add_columns_button_clicked(self):
        pass

    def remove_rows_button_clicked(self):
        pass

    def remove_columns_button_clicked(self):
        pass

  
if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet("QWidget { font-size: 18px; }")  # Apply to all widgets
    window = Main()
    window.show()
    sys.exit(app.exec_())



